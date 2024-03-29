# coding:utf-8
import datetime
from posixpath import split
import pymssql
from bs4 import BeautifulSoup
import sys
import requests
import re
import jieba
import jieba.posseg as pseg
import string
from string import whitespace
nonalpha = string.digits + string.punctuation + string.whitespace
import tkinter as tk
from tkinter import filedialog
from tkinter.filedialog import askopenfilename
import os
import docx2txt
import pdfplumber
import openpyxl
import pandas as pd
import xlsxwriter
from tkinter.filedialog import asksaveasfile

root = tk.Tk()
root.attributes("-topmost", True)
root.withdraw()

#webName = sys.argv[1]
#U_ID = int(sys.argv[2])
#typeID = int(sys.argv[4])
#url = ''

exportFile = filedialog.askopenfilename(title="Open file", filetypes=[("All type","*.html *.htm *.xlsx *.xlsm *.docx *.pdf *.txt")])
a = str(exportFile)
filename, file_extension = os.path.splitext(str(a))
url = file_extension


# html or htm
if (str(file_extension) == '.html' or str(file_extension) == '.htm'):
    soup = BeautifulSoup(open(a, encoding="utf8"), "html.parser")
    tmp = soup.find_all(text=True)

    text = ''
    blacklist = [
        '[document]',
        'noscript',
        'header',
        'html',
        'meta',
        'head', 
        'input',
        'script',
        'style'

    ]
    count = 0
    for t in tmp:
        if t.parent.name not in blacklist:
            tmp = t.text.strip()
            t.string = re.sub(r"[\n][\W]+[^\w]", "\n", tmp)
            text += '{} '.format(tmp)
            count += 1
# text file       
elif (str(file_extension) == '.txt'):
    # print('tes')
    text = ""
    with open(a, encoding="utf-8") as inp:
        for line in inp:
            line = ' '.join(line.split())
            text += line
# word file
elif (str(file_extension) == '.docx'):
    # print("word")
    text = docx2txt.process(a)
    text = ' '.join(text.split())
# pdf file
elif (str(file_extension) == '.pdf'):
    text = ""
    with pdfplumber.open(a) as pdf:
    #Total number of pages
        totalpages = len(pdf.pages)
        for i in range(0 ,totalpages):
            pageobj = pdf.pages[i]
            tmp =  pageobj.extract_text()
            text += ' '.join(tmp.split())
# excel
elif (str(file_extension) == '.xlsx' or str(file_extension) == '.xlsm'):
    dataframe = openpyxl.load_workbook(a)
    dataframe1 = dataframe.active
    
    text = ""
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            if (col[row].value != None):
                text += str(col[row].value) + " "

#decides the flag
def getFirstName(messageContent):
    words = pseg.cut(messageContent)
    for word, flag in words:
        if (flag == 'nr'and len(word) > 3):
            return word

    return False

def getAllName(messageContent):
    words = pseg.cut(messageContent)
    names = []
    for word, flag in words:
        # print('%s,%s' %(word,flag))
        if flag == 'nr':#nr means people's name
            names.append(word)
    return names


def alterWordTagToX(list):
    for x in list:
        jieba.add_word(x, tag='n')

def LoadStopWord():
    # Custom stop words can be added to the list
    StopWordList = [""]

    set(StopWordList)
    alterWordTagToX(StopWordList)

phonePattern = r"[(]?([+]886[-\.\s]?[2-8]|0[2-8])[)]?[-\.\s]?\d{3,4}[-\.\s]?\d{3,4}|(\d{4}|[+]886[-\.\s]?\d{3})[-\.\s]??\d{3}[-\.\s]??\d{3}"

emailPattern = r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+"

def func(value):
    return ''.join(value.splitlines()) # remove space


def extract_address(txt):

    pattern = r"([^\d\s][^\d\s][縣市])[\s]?(\d{1,3}|\d{1,5}|\d{1,3}[-]\d{1,3})?[\s]?(\D{2,9}?(市區|鎮區|鎮市|[鄉鎮市區]))?(\D{2,9}?[路街])[\s]?((\d+?|\D{1,2})[\s]?[段])?[\s]?((\d+?|\D{1,2})[\s]?[巷])?[\s]?(\d+?[弄])?[\s]?((\d+|\D{1,3})[\s]?[號])([\s]?(\d+?|\D{1,3}|\d+(、\d+){0,5})[\s]?[樓])?[\s]?([之][\s]?(\d+?|\D{1,3}))?"
    return re.finditer(pattern, txt, re.MULTILINE)

nameList = []
phoneList = []
mailList = []
addressList = []


for typeID in range(1, 5):
    if (typeID == 1):
        cnt = 0
    
        LoadStopWord()
        final = []
        names = getAllName(text)

        # check if the object is already in list and length is equal to 3
        for n in names:
            if(len(n) == 3 and not n in final):
                final.append(n)

        for x in final:
            nameList.append(x)
        
    elif(typeID == 2):
        cnt = 0

        Empty_list = []
        regex_ex = re.finditer(phonePattern, text, re.MULTILINE)
        for x in regex_ex:
            Empty_list.append(x.group(0))

        y = []
        # check if the object is already in list 
        for x in Empty_list:
            if(x not in y):
                y.append(x)

        for x in y:
            phoneList.append(x)


    elif(typeID == 3):
        cnt = 0

        Empty_list = []
        regex_ex = re.finditer(emailPattern, text, re.MULTILINE)
        for x in regex_ex:
            Empty_list.append(x.group(0))

        y = []
        # check if the object is already in list 
        for x in Empty_list:
            if(x not in y):
                y.append(x)

        for x in y:
            mailList.append(x);

    elif(typeID == 4):
        cnt = 0

        #removes whitespaces and newlines
        out = func(text)
        out.translate(dict.fromkeys(map(ord, whitespace))) 
        out.replace("\\n","")
        out.replace("\\\n","")
        out.replace("\\\\n","")

        res = extract_address(out)


        Empty_list = []
        for x in res:
            Empty_list.append(x.group(0))

        y = []
        # check if the object is already in list and length is greater or equal than 8
        for x in Empty_list:
            if(len(x) >= 8 and x not in y):
                y.append(x.replace(' ', ''))


        for x in y:
            addressList.append(x)


exportFile = filedialog.asksaveasfile(mode='a', title="Save the file", filetypes=[("Excel Workbook", ".xlsx")], initialfile = u"掃描結果", defaultextension=".xlsx")
b = str(exportFile.name)

workbook = xlsxwriter.Workbook(b)
worksheet = workbook.add_worksheet()

List = [nameList, phoneList, mailList, addressList]     

row_num = 0

worksheet.set_column('B:B', 15)
worksheet.set_column('C:C', 30)
worksheet.set_column('D:D', 35)

style = workbook.add_format({'bold': True})
style.set_font_color('#17202A')
style.set_bg_color('#E5E7E9')

description = workbook.add_format({'bold': True})


worksheet.write('A1', '人名', style)
worksheet.write('B1', '電話', style)
worksheet.write('C1', '電子信箱', style)
worksheet.write('D1', '地址', style)

for col_num, data in enumerate(List):
    worksheet.write_column(row_num+1, col_num, data)

workbook.close()